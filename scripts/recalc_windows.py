"""Windows-native variant of the xlsx skill's recalc.py — the shipped script
assumes a Linux sandbox (AF_UNIX shim) which doesn't exist on Windows, so
this skips that shim and calls soffice.exe directly.

WARNING: this REWRITES the target file in place (LibreOffice's store()).
Always point it at a disposable COPY, never at a file you intend to ship —
LibreOffice's OOXML export is not fully Excel-conformant for Tables
combined with shared conditional-formatting styles, and has produced
files that report zero formula errors here but trigger Excel's "repaired
records" dialog on open. Zero errors from this script proves the
formulas are correct; it says nothing about whether the file LibreOffice
just wrote is safe to ship."""
import json
import subprocess
import sys
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

MACRO_FILENAME = "Module1.xba"
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def _stamp(path):
    st = Path(path).stat()
    return st.st_mtime_ns, st.st_size


def recalc(filename, timeout=60):
    abs_path = str(Path(filename).absolute())
    with tempfile.TemporaryDirectory(prefix="recalc-lo-profile-") as profile_dir:
        profile_url = Path(profile_dir).as_uri()

        setup = subprocess.run(
            ["soffice", "--headless", "--terminate_after_init",
             f"-env:UserInstallation={profile_url}"],
            capture_output=True, timeout=timeout,
        )

        macro_dir = Path(profile_dir) / "user" / "basic" / "Standard"
        if not macro_dir.exists():
            return {"error": "LibreOffice did not create a usable profile",
                    "stderr": setup.stderr.decode(errors="replace")}
        (macro_dir / MACRO_FILENAME).write_text(RECALCULATE_MACRO)

        before = _stamp(abs_path)
        cmd = [
            "soffice", "--headless", "--norestore",
            f"-env:UserInstallation={profile_url}",
            "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
            abs_path,
        ]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        except subprocess.TimeoutExpired:
            return {"error": f"LibreOffice timed out after {timeout}s"}

        if result.returncode != 0:
            return {"error": f"soffice exited {result.returncode}: {result.stderr.strip()}"}

        if _stamp(abs_path) == before:
            return {"error": "soffice exited cleanly but never rewrote the file"}

    wb = load_workbook(filename, data_only=True)
    excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
    error_details = {e: [] for e in excel_errors}
    total_errors = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for err in excel_errors:
                        if err in cell.value:
                            error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                            total_errors += 1
                            break
    wb.close()

    wb_f = load_workbook(filename, data_only=False)
    formula_count = 0
    for sheet_name in wb_f.sheetnames:
        for row in wb_f[sheet_name].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, ArrayFormula):
                    v = v.text
                if isinstance(v, str) and v.startswith("="):
                    formula_count += 1
    wb_f.close()

    return {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
        "error_summary": {k: v for k, v in error_details.items() if v},
    }


if __name__ == "__main__":
    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    print(json.dumps(recalc(filename, timeout), indent=2))
