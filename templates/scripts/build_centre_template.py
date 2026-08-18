"""
Generate the Centre Lead data-entry workbook, sharing the same data model
as management/preparation.xlsx: it embeds read-only copies of Management's
reference tables (Centres, Position, ProblemSet, InitiativeType,
AssistanceType, RatingLookup, Resources, Priority, Tactical, Initiative,
Assistance) and adds the Centre Lead's own input tables (Teams, Priorities
& Ranking, Resource Allocation).

Usage:
    python build_centre_template.py <preparation.xlsx> <output.xlsx> [centre-code]
    python build_centre_template.py <preparation.xlsx> --all <output-dir>

The optional centre-code (must match a CentreCode in preparation.xlsx's Centres
table) stamps and locks the Instructions sheet's Centre Name/Code cells at
generation time instead of leaving them as editable placeholders. --all
generates all six centre files in one run, named Centre-<CentreCode>.xlsx.

After running, verify formulas with ../../scripts/recalc_windows.py — but
point it at a COPY of the output, never the file itself. LibreOffice's
store() rewrites the whole file, and its OOXML export is not fully
Excel-conformant for Tables combined with shared conditional-formatting
styles: a LibreOffice-resaved copy of this workbook triggered Excel's
"repaired records" dialog on open even though LibreOffice itself reported
zero formula errors. Ship only the openpyxl-original; use LibreOffice
solely as a disposable formula-correctness check.

Then run ../../management/scripts/wire_data_model.py (with the output
file open in Excel) to wire the tables into the Power Pivot Data Model —
that resave is safe, since it's done by real Excel.
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

FONT = "Arial"
N_ROWS = 200

YELLOW = "FFFFFF00"
GREY = "FFF2F2F2"
HEADER_FILL = "FF1F4E78"
REF_HEADER_FILL = "FF7F7F7F"
HEADER_FONT_COLOR = "FFFFFFFF"
GREEN = "FFC6EFCE"
GREEN_TEXT = "FF006100"
RED = "FFFFC7CE"
RED_TEXT = "FF9C0006"
AMBER = "FFFFEB9C"
AMBER_TEXT = "FF9C6500"

thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

RISK_BANDS = [("Minimal", "FF63BE7B"), ("Low", "FFA9D18E"), ("Moderate", "FFFFEB84"),
              ("High", "FFF4B183"), ("Very High", "FFE06666")]
VALUE_BANDS = [("Minimal", "FFE06666"), ("Limited", "FFF4B183"), ("Moderate", "FFFFEB84"),
               ("Significant", "FFA9D18E"), ("Exceptional", "FF63BE7B")]


def style_header(cell, ref=False):
    cell.font = Font(name=FONT, bold=True, color=HEADER_FONT_COLOR)
    cell.fill = PatternFill("solid", fgColor=REF_HEADER_FILL if ref else HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_body(cell, editable=False, bold=False):
    cell.font = Font(name=FONT, bold=bold)
    cell.border = BORDER
    if editable:
        cell.fill = PatternFill("solid", fgColor=YELLOW)
        cell.protection = Protection(locked=False)


def style_computed(cell):
    """A grey, formula-driven cell: locked (sheet protection must be on
    for that to matter — see input-sheet setup below)."""
    cell.fill = PatternFill("solid", fgColor=GREY)
    cell.font = Font(name=FONT)
    cell.border = BORDER


def read_table(src_wb, sheet_name, header_row=1):
    ws = src_wb[sheet_name]
    headers = [c.value for c in ws[header_row]]
    while headers and headers[-1] is None:
        headers.pop()
    n = len(headers)
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:
            continue
        rows.append(row[:n])
    return headers, rows


def write_reference_table(ws, name, top_row, headers, rows, col_widths=None, start_col=1,
                           protect=True):
    for i, h in enumerate(headers):
        style_header(ws.cell(row=top_row, column=start_col + i, value=h), ref=True)
    for r_off, row in enumerate(rows):
        for c_off, val in enumerate(row):
            style_body(ws.cell(row=top_row + 1 + r_off, column=start_col + c_off, value=val))
    last_row = top_row + len(rows)
    last_col_letter = get_column_letter(start_col + len(headers) - 1)
    first_col_letter = get_column_letter(start_col)
    ref = f"{first_col_letter}{top_row}:{last_col_letter}{last_row}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)
    ws.add_table(tab)
    if col_widths:
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w
    ws.sheet_view.showGridLines = False
    if protect:
        ws.protection.sheet = True
    return last_row


def main(prep_path, out_path, centre_name=None, centre_code=None):
    src = openpyxl.load_workbook(prep_path, data_only=True)

    if centre_code is not None:
        _, centres_rows = read_table(src, "Centres")
        by_code = {row[2]: row[1] for row in centres_rows}  # CentreCode -> Centre
        if centre_code not in by_code:
            raise ValueError(
                f"Unknown centre code {centre_code!r}; must be one of {sorted(by_code)}"
            )
        if centre_name is None:
            centre_name = by_code[centre_code]

    wb = openpyxl.Workbook()

    # ================================================================= Instructions
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    text_blocks = [
        ("Prioritization — Centre Data Entry Template", 14, True),
        ("", None, False),
        ("Purpose", 12, True),
        ("One workbook per centre. Enter this centre's teams, rank the "
         "priorities each team is working on, mark whether each is "
         "resourced, and record how resources split their time across "
         "teams.", None, False),
        ("", None, False),
        ("This workbook shares its data model with Management's Preparation "
         "workbook (data-model/priorities.dbml). The grey-header sheets "
         "(Centres, Position, ProblemSet, InitiativeType, AssistanceType, "
         "RatingLookup, Resources, Priority, Tactical, Initiative, "
         "Assistance) are a READ-ONLY copy of Management's reference data, "
         "protected against editing. The blue-header sheets are yours to "
         "fill in.", None, False),
        ("", None, False),
        ("How to use this workbook", 12, True),
        ("1. 'Step 1 - Teams' — list every team this centre has, whether it "
         "has dedicated resources, and which Priority Type it works on.", None, False),
        ("2. 'Step 2 - Priorities & Ranking' — for each team, pick priorities "
         "from Management's master list (only priorities matching the "
         "team's Priority Type will be offered), rank them, mark whether "
         "the team is actually resourcing it, and what share of the team's "
         "effort it gets.", None, False),
        ("3. 'Step 3 - Resource Allocation' — pick each person from "
         "Management's Resources list, which team they're on, and what "
         "share of their time goes to that team. Their position fills in "
         "automatically.", None, False),
        ("", None, False),
        ("Worked example", 12, True),
        ("Say your centre has a team called 'Analytics Squad', dedicated "
         "(Yes) and working Tactical priorities. On 'Step 1 - Teams' you'd "
         "enter: Team Name = Analytics Squad, Resources Dedicated = Yes, "
         "Priority Type = Tactical. On 'Step 2', a row for that team might "
         "read: Priority Title = (pick one of the Tactical priorities "
         "offered), Rank = 1, Resourced = Yes, Allocation % = 60%. On "
         "'Step 3', a person on that team might read: Resource = (pick "
         "them), Team Name = Analytics Squad, Allocation % = 50%.", None, False),
        ("", None, False),
        ("Legend", 12, True),
        ("Yellow cells = type or pick your data here.", None, False),
        ("Grey 'check' columns = calculated automatically. Don't type in them.", None, False),
        ("Grey-header sheets = read-only reference data from Management.", None, False),
        ("", None, False),
        ("Validation built into this workbook (no macros required)", 12, True),
        ("- Dropdowns keep Team, Type, Priority, Resource, and Position "
         "entries consistent with the master lists, and constrain "
         "Allocation % entries to 5% steps.", None, False),
        ("- The Priority dropdown is filtered to the team's Priority Type "
         "automatically.", None, False),
        ("- Excel will refuse a duplicate team name or a duplicate rank "
         "within a team as you type.", None, False),
        ("- Going over 100% (a team's priorities, or a person's time) isn't "
         "blocked — the dropdown can't check that and still offer a clean "
         "list of 5% steps — but the Total % columns turn red immediately, "
         "so check them before sending this file back.", None, False),
        ("- The duplicate-name/rank checks fire on manual typing. Pasting "
         "many rows at once can bypass them — glance at the check columns "
         "before sending this file back.", None, False),
        ("", None, False),
        ("Centre Info (auto-filled by Management when this workbook was "
         "generated — not something you need to enter)", 12, True),
    ]
    r = 1
    for text, size, bold in text_blocks:
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=FONT, bold=bool(bold), size=size or 11)
        if text and not bold and size is None:
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 30
        r += 1
    centre_name_row = r
    ws.cell(row=r, column=1, value="Centre Name:").font = Font(name=FONT, bold=True)
    style_computed(ws.cell(row=r, column=2, value=centre_name or "Example Centre"))
    r += 1
    centre_code_row = r
    ws.cell(row=r, column=1, value="Centre Code:").font = Font(name=FONT, bold=True)
    style_computed(ws.cell(row=r, column=2, value=centre_code or "EX"))

    wb.defined_names["CentreName"] = DefinedName(
        "CentreName", attr_text=f"Instructions!$B${centre_name_row}")
    wb.defined_names["CentreCode"] = DefinedName(
        "CentreCode", attr_text=f"Instructions!$B${centre_code_row}")
    ws.protection.sheet = True

    # ================================================================= Reference sheets (read-only)
    def copy_ref(sheet_name, col_widths, protect=True):
        headers, rows = read_table(src, sheet_name)
        ws = wb.create_sheet(sheet_name)
        write_reference_table(ws, sheet_name, 1, headers, rows, col_widths, protect=protect)
        return headers, rows

    copy_ref("Centres", [6, 26, 12])
    copy_ref("Position", [26, 14, 10, 8, 12])
    copy_ref("ProblemSet", [6, 28, 14, 18])
    copy_ref("InitiativeType", [6, 32, 14, 18, 18])
    copy_ref("AssistanceType", [6, 28, 14, 18, 18])

    rl_headers, rl_rows = read_table(src, "RatingLookup", header_row=3)
    ws = wb.create_sheet("RatingLookup")
    write_reference_table(ws, "RatingLookup", 1, rl_headers, rl_rows,
                           [6, 14, 10, 10, 14, 12])
    for i, row in enumerate(rl_rows):
        ws.cell(row=2 + i, column=7).fill = PatternFill("solid", fgColor=row[5])
    ws.column_dimensions["G"].width = 4

    res_headers, res_rows = read_table(src, "Resources")
    ws = wb.create_sheet("Resources")
    # FullName is a real formula, not a hardcoded value, and doubles as the
    # picker source for Step 3.
    for i, h in enumerate(res_headers):
        style_header(ws.cell(row=1, column=1 + i, value=h), ref=True)
    style_header(ws.cell(row=1, column=len(res_headers) + 1, value="FullName"), ref=True)
    for r_off, row in enumerate(res_rows):
        row_num = 2 + r_off
        for c_off, val in enumerate(row):
            style_body(ws.cell(row=row_num, column=1 + c_off, value=val))
        style_body(ws.cell(row=row_num, column=len(res_headers) + 1,
                            value=f"=A{row_num}&\" \"&B{row_num}"))
    last_res_row = 1 + len(res_rows)
    tab = Table(displayName="Resources", ref=f"A1:D{last_res_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)
    ws.add_table(tab)
    for col, w in zip("ABCD", [16, 16, 26, 26]):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True
    wb.defined_names["ResourceNameList"] = DefinedName(
        "ResourceNameList", attr_text=f"Resources!$D$2:$D${last_res_row}")

    # Priority: re-sorted so each Type is a contiguous block (needed for the
    # dependent-dropdown named ranges below).
    pr_headers, pr_rows = read_table(src, "Priority")
    type_order = {"Tactical": 0, "Initiative": 1, "Assistance": 2}
    pr_rows_sorted = sorted(pr_rows, key=lambda row: type_order.get(row[1], 99))
    ws = wb.create_sheet("Priority")
    write_reference_table(ws, "Priority", 1, pr_headers, pr_rows_sorted,
                           [42, 12, 10, 12])
    type_bounds = {}
    start = 2
    for t in ("Tactical", "Initiative", "Assistance"):
        count = sum(1 for row in pr_rows_sorted if row[1] == t)
        type_bounds[t] = (start, start + count - 1)
        start += count
    for t, (first, last) in type_bounds.items():
        wb.defined_names[t] = DefinedName(t, attr_text=f"Priority!$A${first}:$A${last}")

    # Table displayName must differ from the sheet name here: the sheet
    # names "Tactical"/"Initiative"/"Assistance" are also used as workbook
    # defined names (for the dependent-dropdown INDIRECT() trick below,
    # via type_bounds). A Table's displayName occupies the same name
    # namespace as a defined name in OOXML — reusing it made Excel's
    # repair silently DELETE the Table (ListObjects came back empty on
    # reopen via COM, confirmed) with no other symptom to point at it.
    for sheet_name, table_name, widths, bands, label_col_letter in (
        ("Tactical", "TacticalScores", [42, 10, 26, 8, 10, 11, 11, 13, 9, 11, 12, 10], RISK_BANDS, "K"),
        ("Initiative", "InitiativeScores", [34, 10, 30, 16, 9, 10, 7, 8, 11, 13, 10], VALUE_BANDS, "J"),
        ("Assistance", "AssistanceScores", [34, 10, 26, 16, 10, 12, 9, 8, 11, 13, 10], VALUE_BANDS, "J"),
    ):
        headers, rows = read_table(src, sheet_name)
        ws = wb.create_sheet(sheet_name)
        last_row = write_reference_table(ws, table_name, 1, headers, rows, widths)
        # FormulaRule, not CellIsRule: a CellIsRule text-equality rule here
        # (structurally valid XML, verified by hand) still made Excel flag
        # these tables for repair on open. FormulaRule is the pattern
        # already proven safe on Step 1's Priority Allocation Total column.
        for band_name, colour in bands:
            ws.conditional_formatting.add(
                f"{label_col_letter}2:{label_col_letter}{last_row}",
                FormulaRule(formula=[f'{label_col_letter}2="{band_name}"'],
                            fill=PatternFill("solid", fgColor=colour))
            )

    # ================================================================= Lookups (small enums, editable-free)
    ws = wb.create_sheet("Lookups")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Small enum lists — used to power dropdowns."
    ws["A1"].font = Font(name=FONT, bold=True)
    ws["A3"] = "Priority Type"
    ws["A3"].font = Font(name=FONT, bold=True)
    for i, val in enumerate(["Tactical", "Initiative", "Assistance"]):
        ws.cell(row=4 + i, column=1, value=val).font = Font(name=FONT)
    wb.defined_names["PriorityTypeList"] = DefinedName(
        "PriorityTypeList", attr_text="Lookups!$A$4:$A$6")
    ws["C3"] = "Resource Status"
    ws["C3"].font = Font(name=FONT, bold=True)
    for i, val in enumerate(["Yes", "No", "Temp"]):
        ws.cell(row=4 + i, column=3, value=val).font = Font(name=FONT)
    wb.defined_names["ResourceStatusList"] = DefinedName(
        "ResourceStatusList", attr_text="Lookups!$C$4:$C$6")
    ws["E3"] = "Yes/No"
    ws["E3"].font = Font(name=FONT, bold=True)
    for i, val in enumerate(["Yes", "No"]):
        ws.cell(row=4 + i, column=5, value=val).font = Font(name=FONT)
    wb.defined_names["YesNoList"] = DefinedName("YesNoList", attr_text="Lookups!$E$4:$E$5")
    ws["G3"] = "Allocation % (5% steps)"
    ws["G3"].font = Font(name=FONT, bold=True)
    for i in range(21):  # 0%, 5%, ..., 100%
        cell = ws.cell(row=4 + i, column=7, value=i * 0.05)
        cell.font = Font(name=FONT)
        cell.number_format = "0%"
    wb.defined_names["PercentIncrementsList"] = DefinedName(
        "PercentIncrementsList", attr_text="Lookups!$G$4:$G$24")
    for col, w in [("A", 16), ("C", 16), ("E", 10), ("G", 20)]:
        ws.column_dimensions[col].width = w

    last_row = N_ROWS + 1

    # ================================================================= Step 1 - Teams
    ws = wb.create_sheet("Step 1 - Teams")
    ws.sheet_view.showGridLines = False
    headers = ["Team Name", "Resources Dedicated", "Priority Type",
               "Priority Allocation Total", "Resource Effort Total (FTE)"]
    widths = [22, 20, 18, 24, 26]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        style_header(ws.cell(row=1, column=i, value=h))

    # Teams starts with exactly one blank data row (not a pre-built block
    # of 200) plus a 14-row buffer below it (Team #2..#15 — CLAUDE.md caps
    # this sheet at 15 teams so it "stays clean"). Only row 2 is part of
    # the Table's official range; rows 3-16 are pre-unlocked, validated,
    # and formatted so typing into row 3 both (a) is actually possible —
    # under sheet protection a locked cell can't be typed into regardless
    # of Table auto-extend, which is what made a true "just one row, grow
    # organically" design impossible — and (b) grows the Table to absorb
    # that row, D/E formulas included, since D/E are genuine calculated
    # columns of this table (not adjacent grey cells outside it, which
    # would silently NOT auto-fill on extend). The small starting size is
    # also what unblocks the Data Model relationship to Teams: Power
    # Pivot rejects a relationship whose "one" side looks like duplicate
    # blanks, which 200 mostly-empty rows guaranteed and 1 does not.
    teams_last_row = 2
    teams_buffer_row = 16
    for r in range(2, teams_buffer_row + 1):
        style_body(ws.cell(row=r, column=1), editable=True)
        style_body(ws.cell(row=r, column=2), editable=True)
        style_body(ws.cell(row=r, column=3), editable=True)
        for col in (4, 5):
            cell = ws.cell(row=r, column=col)
            style_computed(cell)
            cell.number_format = "0%"
        ws.cell(row=r, column=4).value = (
            f'=IF($A{r}="","",SUMIFS(Priorities[Allocation % of Team Effort],'
            f'Priorities[Team Name],$A{r}))'
        )
        ws.cell(row=r, column=5).value = (
            f'=IF($A{r}="","",SUMIFS(ResourceAllocation[Allocation % of Person Effort],'
            f'ResourceAllocation[Team Name],$A{r}))'
        )

    tab = Table(displayName="Teams", ref=f"A1:E{teams_last_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    wb.defined_names["TeamNameList"] = DefinedName("TeamNameList", attr_text="Teams[Team Name]")

    dv_resources = DataValidation(type="list", formula1="=ResourceStatusList", allow_blank=True)
    dv_resources.error = "Choose Yes, No, or Temp."
    dv_resources.errorTitle = "Invalid entry"
    ws.add_data_validation(dv_resources)
    dv_resources.add(f"B2:B{teams_buffer_row}")

    dv_type = DataValidation(type="list", formula1="=PriorityTypeList", allow_blank=True)
    dv_type.error = "Choose Tactical, Initiative, or Assistance."
    dv_type.errorTitle = "Invalid entry"
    ws.add_data_validation(dv_type)
    dv_type.add(f"C2:C{teams_buffer_row}")

    # Plain range, not Teams[Team Name]: a structured reference here (valid
    # syntax, verified by hand) made Excel strip Data Validation from this
    # sheet on open — Data Validation formulas apparently don't support
    # Table structured references the way regular cell formulas do.
    dv_unique_team = DataValidation(
        type="custom",
        formula1=f'=AND($A2<>"",COUNTIF($A$2:$A${teams_buffer_row},$A2)=1)',
        allow_blank=True,
    )
    dv_unique_team.error = "Team names must be unique within this workbook."
    dv_unique_team.errorTitle = "Duplicate team name"
    ws.add_data_validation(dv_unique_team)
    dv_unique_team.add(f"A2:A{teams_buffer_row}")

    ws.protection.sheet = True

    for op, colour, text_colour in (
        ('AND($A2<>"",$D2>1)', RED, RED_TEXT),
        ('AND($A2<>"",$D2<1,$D2>0)', AMBER, AMBER_TEXT),
        ('AND($A2<>"",$D2=1)', GREEN, GREEN_TEXT),
    ):
        ws.conditional_formatting.add(
            f"D2:D{teams_buffer_row}",
            FormulaRule(formula=[op], fill=PatternFill("solid", fgColor=colour),
                        font=Font(name=FONT, color=text_colour))
        )
    ws.freeze_panes = "A2"

    # ================================================================= Step 2 - Priorities & Ranking
    ws = wb.create_sheet("Step 2 - Priorities & Ranking")
    ws.sheet_view.showGridLines = False
    headers = ["Team Name", "Priority Title", "Type (auto)", "Rank", "Resourced",
               "Allocation % of Team Effort"]
    widths = [22, 42, 16, 10, 12, 22]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        style_header(ws.cell(row=1, column=i, value=h))

    for r in range(2, last_row + 1):
        style_body(ws.cell(row=r, column=1), editable=True)
        style_body(ws.cell(row=r, column=2), editable=True)
        c3 = ws.cell(row=r, column=3)
        style_computed(c3)
        c3.value = f'=IF($A{r}="","",IFERROR(INDEX(Teams[Priority Type],MATCH($A{r},Teams[Team Name],0)),"unknown team"))'
        style_body(ws.cell(row=r, column=4), editable=True)
        style_body(ws.cell(row=r, column=5), editable=True)
        style_body(ws.cell(row=r, column=6), editable=True)
        ws.cell(row=r, column=6).number_format = "0%"

    tab = Table(displayName="Priorities", ref=f"A1:F{last_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)

    dv_team2 = DataValidation(type="list", formula1="=TeamNameList", allow_blank=True)
    dv_team2.error = "Pick a team defined on 'Step 1 - Teams'."
    dv_team2.errorTitle = "Unknown team"
    ws.add_data_validation(dv_team2)
    dv_team2.add(f"A2:A{last_row}")

    # Dependent dropdown: the list source is INDIRECT(this row's auto Type),
    # which resolves to the defined name "Tactical"/"Initiative"/"Assistance"
    # — so only priorities matching the team's type are offered.
    dv_priority = DataValidation(type="list", formula1="=INDIRECT($C2)", allow_blank=True)
    dv_priority.error = "Pick a priority from Management's master list, matching the team's Priority Type."
    dv_priority.errorTitle = "Unknown priority"
    ws.add_data_validation(dv_priority)
    dv_priority.add(f"B2:B{last_row}")

    dv_resourced = DataValidation(type="list", formula1="=YesNoList", allow_blank=True)
    dv_resourced.error = "Choose Yes or No."
    dv_resourced.errorTitle = "Invalid entry"
    ws.add_data_validation(dv_resourced)
    dv_resourced.add(f"E2:E{last_row}")

    # Plain ranges here too — see the note on dv_unique_team above.
    dv_rank = DataValidation(
        type="custom",
        formula1=(f'=AND($D2<>"",$D2=INT($D2),$D2>0,'
                  f'COUNTIFS($A$2:$A${last_row},$A2,$D$2:$D${last_row},$D2)=1)'),
        allow_blank=True,
    )
    dv_rank.error = "Rank must be a positive whole number, unique within the team."
    dv_rank.errorTitle = "Invalid rank"
    ws.add_data_validation(dv_rank)
    dv_rank.add(f"D2:D{last_row}")

    # A dropdown (list) and a hard "team total <=100%" block can't coexist
    # in one native Data Validation rule — this picks the dropdown; the
    # Priority Allocation Total column on Step 1 still flags an overage
    # visually (red), just not as a typing-time block.
    dv_alloc2 = DataValidation(type="list", formula1="=PercentIncrementsList", allow_blank=True)
    dv_alloc2.error = "Pick a value from the dropdown (5% steps)."
    dv_alloc2.errorTitle = "Invalid entry"
    ws.add_data_validation(dv_alloc2)
    dv_alloc2.add(f"F2:F{last_row}")

    ws.protection.sheet = True
    ws.freeze_panes = "A2"

    # ================================================================= Step 3 - Resource Allocation
    ws = wb.create_sheet("Step 3 - Resource Allocation")
    ws.sheet_view.showGridLines = False
    headers = ["Resource", "Position Title (auto)", "Team Name",
               "Allocation % of Person Effort", "Person Total %"]
    widths = [26, 26, 22, 24, 16]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        style_header(ws.cell(row=1, column=i, value=h))

    for r in range(2, last_row + 1):
        style_body(ws.cell(row=r, column=1), editable=True)
        c2 = ws.cell(row=r, column=2)
        style_computed(c2)
        c2.value = f'=IF($A{r}="","",IFERROR(INDEX(Resources[PositionTitle],MATCH($A{r},Resources[FullName],0)),"unknown resource"))'
        style_body(ws.cell(row=r, column=3), editable=True)
        style_body(ws.cell(row=r, column=4), editable=True)
        ws.cell(row=r, column=4).number_format = "0%"
        c5 = ws.cell(row=r, column=5)
        style_computed(c5)
        c5.number_format = "0%"
        c5.value = (
            f'=IF($A{r}="","",SUMIFS(ResourceAllocation[Allocation % of Person Effort],'
            f'ResourceAllocation[Resource],$A{r}))'
        )

    tab = Table(displayName="ResourceAllocation", ref=f"A1:E{last_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)

    dv_res_name = DataValidation(type="list", formula1="=ResourceNameList", allow_blank=True)
    dv_res_name.error = "Pick a person from Management's Resources list."
    dv_res_name.errorTitle = "Unknown resource"
    ws.add_data_validation(dv_res_name)
    dv_res_name.add(f"A2:A{last_row}")

    dv_team3 = DataValidation(type="list", formula1="=TeamNameList", allow_blank=True)
    dv_team3.error = "Pick a team defined on 'Step 1 - Teams'."
    dv_team3.errorTitle = "Unknown team"
    ws.add_data_validation(dv_team3)
    dv_team3.add(f"C2:C{last_row}")

    # Dropdown, not a hard block — see the note on dv_alloc2 in Step 2.
    # Person Total % still flags an overage visually (red).
    dv_alloc3 = DataValidation(type="list", formula1="=PercentIncrementsList", allow_blank=True)
    dv_alloc3.error = "Pick a value from the dropdown (5% steps)."
    dv_alloc3.errorTitle = "Invalid entry"
    ws.add_data_validation(dv_alloc3)
    dv_alloc3.add(f"D2:D{last_row}")

    ws.protection.sheet = True

    for op, colour, text_colour in (
        ('AND($A2<>"",$E2>1)', RED, RED_TEXT),
        ('AND($A2<>"",$E2=1)', GREEN, GREEN_TEXT),
    ):
        ws.conditional_formatting.add(
            f"E2:E{last_row}",
            FormulaRule(formula=[op], fill=PatternFill("solid", fgColor=colour),
                        font=Font(name=FONT, color=text_colour))
        )
    ws.freeze_panes = "A2"

    wb.move_sheet("Instructions", offset=-20)
    wb.active = 0
    wb.save(out_path)
    print("saved", out_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(
            "Usage:\n"
            "  python build_centre_template.py <preparation.xlsx> <output.xlsx> [centre-code]\n"
            "  python build_centre_template.py <preparation.xlsx> --all <output-dir>"
        )
    prep_arg = sys.argv[1]
    if sys.argv[2] == "--all":
        if len(sys.argv) != 4:
            raise SystemExit("Usage: python build_centre_template.py <preparation.xlsx> --all <output-dir>")
        out_dir = Path(sys.argv[3])
        out_dir.mkdir(parents=True, exist_ok=True)
        centres_src = openpyxl.load_workbook(prep_arg, data_only=True)
        _, all_centres = read_table(centres_src, "Centres")
        for _id, c_name, c_code in all_centres:
            main(prep_arg, str(out_dir / f"Centre-{c_code}.xlsx"), centre_name=c_name, centre_code=c_code)
    else:
        main(prep_arg, sys.argv[2], centre_code=sys.argv[3] if len(sys.argv) > 3 else None)
