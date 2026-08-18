"""
Dev-only fixture generator: produces a handful of small, hand-checkable
"completed" centre workbooks (Teams/Priorities/ResourceAllocation filled in)
for testing the Consolidation workbook against — nothing like this existed
anywhere in the repo, since real Centre Lead submissions don't exist yet.

Output goes to dev_fixtures/output/ (gitignored) — these are throwaway,
regenerate any time by re-running this script. Not a shipped deliverable.

Reuses build_centre_template.main() (so fixtures go through the real,
current template-generation code path, including the Centre Name/Code
stamping) and read_table() to pull real Priority/Resource names out of
preparation.xlsx rather than inventing fake ones.

Deliberately plants a Team Name collision across two centres (both ECO and
INF have a team called "Ops Team") — this is the exact cross-centre
collision the Consolidation workbook's TeamKey (CentreCode | Team Name)
composite key exists to handle. If TeamKey wiring is ever wrong, ECO's and
INF's "Ops Team" rows will get merged and their FTE sums will be wrong (see
expected values below).

Usage:
    python make_test_centres.py <preparation.xlsx>

Then, with each output file open in Excel:
    python ../../../templates/scripts/wire_data_model.py Centre-CYB.xlsx
    (etc. for ECO, INF)
    python ../../../scripts/check_errors_excel.py Centre-CYB.xlsx

Expected hand-computed values (Allocation % of Person Effort := SUM per team;
FTE Delivered to Priority := that sum x Allocation % of Team Effort):

    CYB  Firewall Team FTE = 0.50 + 0.50 = 1.00
         -> Harden network perimeter defenses:  1.00 x 0.60 = 0.60
         -> Close critical patching gap:        1.00 x 0.40 = 0.40
    CYB  Threat Intel FTE  = 1.00
         -> Launch predictive analytics pilot:  1.00 x 1.00 = 1.00
    ECO  Ops Team FTE = 0.75
         -> Provide cross-centre threat-briefing support: 0.75 x 1.00 = 0.75
    INF  Ops Team FTE = 0.25   (a DIFFERENT team from ECO's "Ops Team" —
         this is the collision case)
         -> Reduce third-party vendor risk exposure: 0.25 x 1.00 = 0.25

    Total FTE across all three fixture centres = 3.00
    (If ECO/INF's "Ops Team" rows ever get merged by Team Name instead of
    TeamKey, you'd wrongly see a single 1.00-FTE "Ops Team" instead of the
    correct separate 0.75 and 0.25 — that's the regression this fixture
    set is designed to catch.)
"""
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "templates" / "scripts"))
from build_centre_template import main as build_centre, read_table  # noqa: E402

OUT_DIR = Path(__file__).resolve().parent / "output"

# (centre_code, teams, priorities, resource_allocations)
#   teams:        [(team_name, resources_dedicated, priority_type), ...]
#   priorities:   [(team_name, priority_title, rank, resourced, alloc_pct), ...]
#   allocations:  [(resource_full_name, team_name, alloc_pct), ...]
FIXTURES = [
    (
        "CYB",
        [
            ("Firewall Team", "Yes", "Tactical"),
            ("Threat Intel", "Yes", "Initiative"),
        ],
        [
            ("Firewall Team", "Harden network perimeter defenses", 1, "Yes", 0.60),
            ("Firewall Team", "Close critical patching gap", 2, "Yes", 0.40),
            ("Threat Intel", "Launch predictive analytics pilot", 1, "Yes", 1.00),
        ],
        [
            ("Kevin Johnson", "Firewall Team", 0.50),
            ("John Brown", "Firewall Team", 0.50),
            ("Chen Johnson", "Threat Intel", 1.00),
        ],
    ),
    (
        "ECO",
        [
            ("Ops Team", "Yes", "Assistance"),
        ],
        [
            ("Ops Team", "Provide cross-centre threat-briefing support", 1, "Yes", 1.00),
        ],
        [
            ("Thomas Kowalski", "Ops Team", 0.75),
        ],
    ),
    (
        "INF",
        [
            ("Ops Team", "Yes", "Tactical"),  # same name as ECO's team, different centre
        ],
        [
            ("Ops Team", "Reduce third-party vendor risk exposure", 1, "Yes", 1.00),
        ],
        [
            ("Emma Chen", "Ops Team", 0.25),
        ],
    ),
]


def fill_fixture(out_path, teams, priorities, allocations):
    wb = openpyxl.load_workbook(out_path)

    ws = wb["Step 1 - Teams"]
    for i, (name, dedicated, ptype) in enumerate(teams):
        row = 2 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=dedicated)
        ws.cell(row=row, column=3, value=ptype)
    last_team_row = 1 + len(teams)
    ws.tables["Teams"].ref = f"A1:E{last_team_row}"

    ws = wb["Step 2 - Priorities & Ranking"]
    for i, (team, title, rank, resourced, pct) in enumerate(priorities):
        row = 2 + i
        ws.cell(row=row, column=1, value=team)
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=4, value=rank)
        ws.cell(row=row, column=5, value=resourced)
        cell = ws.cell(row=row, column=6, value=pct)
        cell.number_format = "0%"

    ws = wb["Step 3 - Resource Allocation"]
    for i, (resource, team, pct) in enumerate(allocations):
        row = 2 + i
        ws.cell(row=row, column=1, value=resource)
        ws.cell(row=row, column=3, value=team)
        cell = ws.cell(row=row, column=4, value=pct)
        cell.number_format = "0%"

    wb.save(out_path)


def main(prep_path):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    _, centres_rows = read_table(openpyxl.load_workbook(prep_path, data_only=True), "Centres")
    names_by_code = {row[2]: row[1] for row in centres_rows}

    for code, teams, priorities, allocations in FIXTURES:
        out_path = OUT_DIR / f"Centre-{code}.xlsx"
        build_centre(prep_path, str(out_path), centre_name=names_by_code[code], centre_code=code)
        fill_fixture(out_path, teams, priorities, allocations)
        print(f"fixture ready: {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python make_test_centres.py <preparation.xlsx>")
    main(sys.argv[1])
