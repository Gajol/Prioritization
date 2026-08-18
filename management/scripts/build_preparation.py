"""
Generate the Management Preparation workbook: reference/master tables
(Centres, Position, ProblemSet, InitiativeType, AssistanceType,
RatingLookup) plus synthesized test data (100 Resources, 50 Priorities
with full Tactical/Initiative/Assistance scoring), per
data-model/priorities.dbml.

Usage:
    python build_preparation.py <output.xlsx>

After running, verify formulas with ../../scripts/recalc_windows.py — but
point it at a COPY of the output, never the file itself. LibreOffice's
store() rewrites the whole file, and its OOXML export is not fully
Excel-conformant for Tables combined with shared conditional-formatting
styles (this shipped a "repaired records" file once already — see that
script's own history). Ship only the openpyxl-original; use LibreOffice
solely as a disposable formula-correctness check.

Then run wire_data_model.py (with the output file open in Excel) to wire
the tables into the Power Pivot Data Model — that resave is safe, since
it's done by real Excel.
"""
import random
import sys

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

random.seed(42)

FONT = "Arial"
HEADER_FILL = "FF1F4E78"
HEADER_FONT_COLOR = "FFFFFFFF"
GREY = "FFF2F2F2"

thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell):
    cell.font = Font(name=FONT, bold=True, color=HEADER_FONT_COLOR)
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_body(cell, bold=False):
    cell.font = Font(name=FONT, bold=bold)
    cell.border = BORDER


def write_table(ws, wb, name, top_row, headers, rows, col_widths=None, start_col=1):
    for i, h in enumerate(headers):
        style_header(ws.cell(row=top_row, column=start_col + i, value=h))
    for r_off, row in enumerate(rows):
        for c_off, val in enumerate(row):
            style_body(ws.cell(row=top_row + 1 + r_off, column=start_col + c_off, value=val))
    last_row = top_row + len(rows)
    last_col_letter = get_column_letter(start_col + len(headers) - 1)
    first_col_letter = get_column_letter(start_col)
    ref = f"{first_col_letter}{top_row}:{last_col_letter}{last_row}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    if col_widths:
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w
    return last_row


wb = openpyxl.Workbook()

# =====================================================================
# Instructions
# =====================================================================
ws = wb.active
ws.title = "Instructions"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 100

text_blocks = [
    ("Prioritization — Management Preparation Workbook", 14, True),
    ("", None, False),
    ("Purpose", 12, True),
    ("This workbook is Management's Preparation-phase deliverable. It holds "
     "the reference/master data every Centre Lead workbook depends on: the "
     "list of Centres, the Position roster, the master Priority list with "
     "full Tactical/Initiative/Assistance scoring, and the RatingLookup "
     "bands used to label and colour-code scores.", None, False),
    ("", None, False),
    ("It uses the SAME data model as the Centre Lead workbooks (same table "
     "names/columns, defined from data-model/priorities.dbml). Centre Leads "
     "receive copies of the reference tables here (Centres, Position, "
     "Resources, Priority, Tactical, Initiative, Assistance, RatingLookup) "
     "plus their own blank transactional tables (Teams, TeamPriorities, "
     "TeamPriorityAllocation, ResourceAllocation).", None, False),
    ("", None, False),
    ("Sheets", 12, True),
    ("Centres, Position, ProblemSet, InitiativeType, AssistanceType — small "
     "reference lookups.", None, False),
    ("RatingLookup — band definitions (name + colour) for Likelihood, Risk, "
     "and Value scores. See 'Simplifications' below.", None, False),
    ("Resources — 100 synthesized people across the Position roster.", None, False),
    ("Priority — 50 synthesized priorities split across the three types.", None, False),
    ("Tactical / Initiative / Assistance — one scoring row per matching "
     "Priority, with computed Likelihood/Risk or Value and a band label + "
     "colour looked up from RatingLookup.", None, False),
    ("", None, False),
    ("All data on every sheet except Instructions is SYNTHESIZED for design "
     "and testing purposes — no real people, centres, or priorities.", None, True),
    ("", None, False),
    ("Simplifications from the draft data model (data-model/priorities.dbml)", 12, True),
    ("1. RatingLookup bands: the six raw 0-5 inputs (Intent, Capability, "
     "Consequence, Dividend, Feasibility, Cost) are left as plain numbers — "
     "only the computed outcomes (Likelihood, Risk, Value) are banded and "
     "colour-coded, since those are what get reviewed/visualized.", None, False),
    ("2. RiskBandLevel is computed as LikelihoodLevel (the 1-5 band index of "
     "the raw Likelihood score) x Consequence, giving a 0-25 range — the "
     "draft model's wording ('Likelihood * Consequence') was ambiguous "
     "between the raw 0-25 Likelihood and its banded 1-5 level; the banded "
     "version keeps RiskBandLevel on a comparable 0-25 scale.", None, False),
    ("3. Actor now stores a Centre Code (FK to Centres) instead of the "
     "draft's fixed Ottawa/Gatineau/Toronto enum, so it works for all six "
     "centres.", None, False),
    ("4. LikelihoodScore and LikelihoodLevel were both defined in the draft "
     "with the same meaning — this workbook keeps only LikelihoodLevel.", None, False),
    ("", None, False),
    ("Next step: wire these tables into Excel's Data Model (Power Pivot) "
     "with relationships matching the Ref: lines in the DBML, so Centre "
     "workbooks can roll up through it. See docs/ once written.", None, True),
]

r = 1
for text, size, bold in text_blocks:
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name=FONT, bold=bool(bold), size=size or 11)
    if text and not bold and size is None:
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 30
    r += 1

# =====================================================================
# Centres
# =====================================================================
ws = wb.create_sheet("Centres")
ws.sheet_view.showGridLines = False
centres = [
    (1, "Cyber Centre", "CYB"),
    (2, "Economic Centre", "ECO"),
    (3, "Infrastructure Centre", "INF"),
    (4, "Health Centre", "HLT"),
    (5, "Environmental Centre", "ENV"),
    (6, "Diplomatic Centre", "DIP"),
]
write_table(ws, wb, "Centres", 1, ["id", "Centre", "CentreCode"], centres,
            col_widths=[6, 26, 12])
CENTRE_CODES = [c[2] for c in centres]

# =====================================================================
# Position
# =====================================================================
ws = wb.create_sheet("Position")
ws.sheet_view.showGridLines = False
positions = [
    ("Operations Analyst", "Operations", "OA", "L1", 1.0),
    ("Senior Operations Analyst", "Operations", "OA", "L2", 2.0),
    ("Operations Manager", "Operations", "OM", "L3", 3.0),
    ("Senior Operations Manager", "Operations", "OM", "L4", 4.0),
    ("Director of Operations", "Operations", "DIR", "L5", 5.0),
    ("Technical Analyst", "Technical", "TA", "L1", 1.0),
    ("Software Engineer", "Technical", "SWE", "L2", 2.0),
    ("Senior Software Engineer", "Technical", "SWE", "L3", 3.0),
    ("Technical Lead", "Technical", "TL", "L4", 4.0),
    ("Director of Technology", "Technical", "DIR", "L5", 5.0),
    ("Legal Counsel", "Legal", "LC", "L3", 3.0),
    ("Senior Legal Counsel", "Legal", "LC", "L4", 4.0),
    ("General Counsel", "Legal", "GC", "L5", 5.0),
]
write_table(ws, wb, "Position", 1,
            ["Title", "PositionArea", "Position", "Level", "SortingHelper"],
            positions, col_widths=[26, 14, 10, 8, 12])
POSITION_TITLES = [p[0] for p in positions]
POSITION_WEIGHTS = [3, 2, 3, 2, 1, 3, 3, 2, 2, 1, 2, 1, 1]  # more junior roles

# =====================================================================
# ProblemSet
# =====================================================================
ws = wb.create_sheet("ProblemSet")
ws.sheet_view.showGridLines = False
problem_sets = [
    (1, "Supply Chain Resilience", "Efficiency", "SCR"),
    (2, "Revenue Diversification", "Profitability", "REV"),
    (3, "Legacy System Modernization", "Efficiency", "LSM"),
    (4, "Fraud Reduction", "Profitability", "FRD"),
    (5, "Talent Retention", "Efficiency", "TAL"),
]
write_table(ws, wb, "ProblemSet", 1, ["id", "Title", "TacticalType", "ProblemSetShortCode"],
            problem_sets, col_widths=[6, 28, 14, 18])

# =====================================================================
# InitiativeType
# =====================================================================
ws = wb.create_sheet("InitiativeType")
ws.sheet_view.showGridLines = False
initiative_types = [
    (1, "Data Exploitation Initiative", "Analytical", "DataExploitation", "DEI"),
    (2, "Technical Discovery Initiative", "Discovery", "Technical", "TDI"),
    (3, "Partnership Coordination Initiative", "Partnership", "Coordination", "PCI"),
    (4, "Analytical Partnership Initiative", "Partnership", "DataExploitation", "API"),
]
write_table(ws, wb, "InitiativeType", 1,
            ["id", "Title", "Level1", "Level2", "InitiativeShortCode"],
            initiative_types, col_widths=[6, 32, 14, 18, 18])

# =====================================================================
# AssistanceType
# =====================================================================
ws = wb.create_sheet("AssistanceType")
ws.sheet_view.showGridLines = False
assistance_types = [
    (1, "Advisory Support", "Guidance", "Advisory", "ADV"),
    (2, "Technical Assistance", "Guidance", "Technical", "TEC"),
    (3, "Capacity Building", "Development", "Training", "CAP"),
    (4, "Liaison Support", "Development", "Coordination", "LIA"),
]
write_table(ws, wb, "AssistanceType", 1,
            ["id", "Title", "Level1", "Level2", "AssistanceShortCode"],
            assistance_types, col_widths=[6, 28, 14, 18, 18])

# =====================================================================
# RatingLookup
# =====================================================================
ws = wb.create_sheet("RatingLookup")
ws.sheet_view.showGridLines = False
ws["A1"] = ("Band definitions for computed scores. Six raw 0-5 inputs are left "
            "unlabelled by design — see Instructions, Simplification #1.")
ws["A1"].font = Font(name=FONT, italic=True)
ws.merge_cells("A1:F1")

rating_rows = [
    # id, RatingType, minValue, maxValue, BandName, ColourCode
    (1, "Likelihood", 0, 5, "Very Low", "FFDCE6F1"),
    (2, "Likelihood", 5, 10, "Low", "FFB8CCE4"),
    (3, "Likelihood", 10, 15, "Moderate", "FF8DB4E2"),
    (4, "Likelihood", 15, 20, "High", "FF538DD5"),
    (5, "Likelihood", 20, 25, "Very High", "FF1F497D"),
    (6, "Risk", 0, 5, "Minimal", "FF63BE7B"),
    (7, "Risk", 5, 10, "Low", "FFA9D18E"),
    (8, "Risk", 10, 15, "Moderate", "FFFFEB84"),
    (9, "Risk", 15, 20, "High", "FFF4B183"),
    (10, "Risk", 20, 25, "Very High", "FFE06666"),
    (11, "Value", 0, 3, "Minimal", "FFE06666"),
    (12, "Value", 3, 6, "Limited", "FFF4B183"),
    (13, "Value", 6, 9, "Moderate", "FFFFEB84"),
    (14, "Value", 9, 12, "Significant", "FFA9D18E"),
    (15, "Value", 12, 15, "Exceptional", "FF63BE7B"),
]
last_rl = write_table(ws, wb, "RatingLookup", 3,
                       ["id", "RatingType", "minValue", "maxValue", "BandName", "ColourCode"],
                       rating_rows, col_widths=[6, 14, 10, 10, 14, 12])
# swatch preview
for i, row in enumerate(rating_rows):
    ws.cell(row=4 + i, column=7).fill = PatternFill("solid", fgColor=row[5])
ws.column_dimensions["G"].width = 4

# Per-type named ranges (5 contiguous rows each) power the INDEX/MATCH band
# lookups on Tactical/Initiative/Assistance without needing array formulas.
def add_band_names(prefix, start_row):
    wb.defined_names[f"{prefix}_Min"] = DefinedName(
        f"{prefix}_Min", attr_text=f"RatingLookup!$C${start_row}:$C${start_row + 4}")
    wb.defined_names[f"{prefix}_Name"] = DefinedName(
        f"{prefix}_Name", attr_text=f"RatingLookup!$E${start_row}:$E${start_row + 4}")
    wb.defined_names[f"{prefix}_Colour"] = DefinedName(
        f"{prefix}_Colour", attr_text=f"RatingLookup!$F${start_row}:$F${start_row + 4}")
    wb.defined_names[f"{prefix}_Id"] = DefinedName(
        f"{prefix}_Id", attr_text=f"RatingLookup!$A${start_row}:$A${start_row + 4}")

add_band_names("Likelihood", 4)
add_band_names("Risk", 9)
add_band_names("Value", 14)

# =====================================================================
# Resources (100 synthesized people)
# =====================================================================
FIRST_NAMES = ["James", "Maria", "David", "Sarah", "Michael", "Priya", "Wei", "Fatima",
               "John", "Anna", "Carlos", "Aisha", "Robert", "Emma", "Ahmed", "Sophie",
               "Daniel", "Yuki", "Omar", "Laura", "Kevin", "Nadia", "Peter", "Chen",
               "Ibrahim", "Grace", "Marc", "Leila", "Thomas", "Elena"]
LAST_NAMES = ["Smith", "Chen", "Patel", "Johnson", "Garcia", "Kim", "Nguyen", "Brown",
              "Singh", "Martin", "Lee", "Khan", "Wilson", "Rossi", "Dubois", "Andersson",
              "Silva", "Kowalski", "Yamamoto", "Novak", "Cohen", "Murphy", "Haddad",
              "Ivanov", "Costa", "Larsen", "Petit", "Osei", "Tanaka", "Fischer"]

used_names = set()
resources = []
while len(resources) < 100:
    fn = random.choice(FIRST_NAMES)
    ln = random.choice(LAST_NAMES)
    if (fn, ln) in used_names:
        continue
    used_names.add((fn, ln))
    pos = random.choices(POSITION_TITLES, weights=POSITION_WEIGHTS, k=1)[0]
    resources.append((fn, ln, pos))

ws = wb.create_sheet("Resources")
ws.sheet_view.showGridLines = False
write_table(ws, wb, "Resources", 1, ["FirstName", "LastName", "PositionTitle"], resources,
            col_widths=[16, 16, 26])

# =====================================================================
# Priority (50 synthesized, split across the three types)
# =====================================================================
TACTICAL_TITLES = [
    "Harden network perimeter defenses",
    "Close critical patching gap",
    "Reduce third-party vendor risk exposure",
    "Contain phishing incident recurrence",
    "Stabilize legacy authentication systems",
    "Mitigate single-vendor supply dependency",
    "Address currency exposure on cross-border contracts",
    "Reduce fraud loss in claims processing",
    "Shore up disaster-recovery failover gap",
    "Contain cost overruns on flagship program",
    "Close skills gap in incident response team",
    "Reduce backlog of unresolved audit findings",
    "Mitigate insider-threat detection blind spot",
    "Stabilize vendor payment processing errors",
    "Address data-retention compliance gap",
    "Reduce downtime on critical case-management platform",
    "Contain reputational risk from delayed public reporting",
]

INITIATIVE_TITLES = [
    "Launch predictive analytics pilot",
    "Modernize case-management platform",
    "Stand up cross-centre data-sharing hub",
    "Build automated reporting dashboards",
    "Pilot AI-assisted triage for intake requests",
    "Develop centre-wide skills academy",
    "Launch digital self-service portal",
    "Establish innovation partnership with academia",
    "Build a unified risk-scoring model",
    "Modernize records-management architecture",
    "Pilot flexible/hybrid staffing model",
    "Develop a real-time performance dashboard",
    "Launch a data-quality improvement program",
    "Build a shared services centre for back-office functions",
    "Pilot a client-experience redesign",
    "Develop an early-warning indicators program",
]

ASSISTANCE_TITLES = [
    "Provide cross-centre threat-briefing support",
    "Second an analyst to a partner centre",
    "Share technical infrastructure capacity",
    "Provide legal review support to partner initiatives",
    "Coordinate joint training exercises",
    "Provide surge staffing during peak demand",
    "Share lessons learned from recent incidents",
    "Provide advisory support on procurement",
    "Coordinate a joint public-communications response",
    "Provide mentoring capacity to junior analysts",
    "Share data infrastructure with partner centre",
    "Provide translation/interpretation support",
    "Coordinate a joint tabletop exercise",
    "Provide technical assistance on system migration",
    "Support a partner centre's audit response",
    "Provide surge analytic capacity for a major event",
]

assert len(TACTICAL_TITLES) + len(INITIATIVE_TITLES) + len(ASSISTANCE_TITLES) == 49

priorities = []
for title in TACTICAL_TITLES:
    priorities.append((title, "Tactical", "Risk", "No"))
for title in INITIATIVE_TITLES:
    priorities.append((title, "Initiative", "Value", "No"))
for title in ASSISTANCE_TITLES:
    priorities.append((title, "Assistance", "Value", "No"))
# round out to an even 50
priorities.append(("Modernize vendor risk-assessment process", "Tactical", "Risk", "No"))

ws = wb.create_sheet("Priority")
ws.sheet_view.showGridLines = False
write_table(ws, wb, "Priority", 1, ["Title", "Type", "Impact", "Resources"], priorities,
            col_widths=[42, 12, 10, 12])
wb.defined_names["PriorityTitleList"] = DefinedName(
    "PriorityTitleList", attr_text=f"Priority!$A$2:$A${1 + len(priorities)}")

SCORE_WEIGHTS = [1, 3, 4, 4, 3, 1]  # bell-ish curve over 0..5
def score():
    return random.choices([0, 1, 2, 3, 4, 5], weights=SCORE_WEIGHTS, k=1)[0]


def band_formulas(raw_cell, id_col_letter, row, prefix):
    """Return (id_formula, name_formula, colour_formula) for a raw score cell."""
    id_f = (f'=INDEX({prefix}_Id,MATCH({raw_cell},{prefix}_Min,1))')
    name_f = f'=INDEX({prefix}_Name,MATCH({id_col_letter}{row},{prefix}_Id,0))'
    colour_f = f'=INDEX({prefix}_Colour,MATCH({id_col_letter}{row},{prefix}_Id,0))'
    return id_f, name_f, colour_f


# ---------------------------------------------------------------- Tactical
ws = wb.create_sheet("Tactical")
ws.sheet_view.showGridLines = False
headers = ["PriorityReference", "Actor", "ProblemSet", "Intent", "Capability",
           "Consequence", "Likelihood", "LikelihoodLevel", "RiskRaw",
           "RiskLabelId", "RiskLabel", "RiskColour"]
widths = [42, 10, 26, 8, 10, 11, 11, 13, 9, 11, 12, 10]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws.cell(row=1, column=i, value=h))

tactical_rows = [p for p in priorities if p[1] == "Tactical"]
r = 2
for title, _, _, _ in tactical_rows:
    intent, capability, consequence = score(), score(), score()
    actor = random.choice(CENTRE_CODES)
    pset = random.choice(problem_sets)[1]
    style_body(ws.cell(row=r, column=1, value=title))
    style_body(ws.cell(row=r, column=2, value=actor))
    style_body(ws.cell(row=r, column=3, value=pset))
    style_body(ws.cell(row=r, column=4, value=intent))
    style_body(ws.cell(row=r, column=5, value=capability))
    style_body(ws.cell(row=r, column=6, value=consequence))
    style_body(ws.cell(row=r, column=7, value=f"=D{r}*E{r}"))
    style_body(ws.cell(row=r, column=8, value=f"=INDEX(Likelihood_Id,MATCH(G{r},Likelihood_Min,1))"))
    style_body(ws.cell(row=r, column=9, value=f"=H{r}*F{r}"))
    id_f, name_f, colour_f = band_formulas("I" + str(r), "J", r, "Risk")
    style_body(ws.cell(row=r, column=10, value=id_f))
    c11 = ws.cell(row=r, column=11, value=name_f)
    style_body(c11)
    c12 = ws.cell(row=r, column=12, value=colour_f)
    style_body(c12)
    r += 1
last_tactical = r - 1
tab = Table(displayName="Tactical", ref=f"A1:L{last_tactical}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = "A2"

# ---------------------------------------------------------------- Initiative
ws = wb.create_sheet("Initiative")
ws.sheet_view.showGridLines = False
headers = ["PriorityReference", "Actor", "InitiativeType", "InitiativeShortCode",
           "Dividend", "Feasibility", "Cost", "Value", "ValueLabelId", "ValueLabel",
           "ValueColour"]
widths = [34, 10, 30, 16, 9, 10, 7, 8, 11, 13, 10]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws.cell(row=1, column=i, value=h))

initiative_rows = [p for p in priorities if p[1] == "Initiative"]
r = 2
for title, _, _, _ in initiative_rows:
    dividend, feasibility, cost = score(), score(), score()
    actor = random.choice(CENTRE_CODES)
    itype = random.choice(initiative_types)
    style_body(ws.cell(row=r, column=1, value=title))
    style_body(ws.cell(row=r, column=2, value=actor))
    style_body(ws.cell(row=r, column=3, value=itype[1]))
    style_body(ws.cell(row=r, column=4, value=itype[4]))
    style_body(ws.cell(row=r, column=5, value=dividend))
    style_body(ws.cell(row=r, column=6, value=feasibility))
    style_body(ws.cell(row=r, column=7, value=cost))
    style_body(ws.cell(row=r, column=8, value=f"=E{r}+F{r}+G{r}"))
    id_f, name_f, colour_f = band_formulas("H" + str(r), "I", r, "Value")
    style_body(ws.cell(row=r, column=9, value=id_f))
    style_body(ws.cell(row=r, column=10, value=name_f))
    style_body(ws.cell(row=r, column=11, value=colour_f))
    r += 1
last_initiative = r - 1
tab = Table(displayName="Initiative", ref=f"A1:K{last_initiative}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = "A2"

# ---------------------------------------------------------------- Assistance
ws = wb.create_sheet("Assistance")
ws.sheet_view.showGridLines = False
headers = ["PriorityReference", "Actor", "AssistanceType", "AssistanceShortCode",
           "Alignment", "Contribution", "Capacity", "Value", "ValueLabelId",
           "ValueLabel", "ValueColour"]
widths = [34, 10, 26, 16, 10, 12, 9, 8, 11, 13, 10]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws.cell(row=1, column=i, value=h))

assistance_rows = [p for p in priorities if p[1] == "Assistance"]
r = 2
for title, _, _, _ in assistance_rows:
    alignment, contribution, capacity = score(), score(), score()
    actor = random.choice(CENTRE_CODES)
    atype = random.choice(assistance_types)
    style_body(ws.cell(row=r, column=1, value=title))
    style_body(ws.cell(row=r, column=2, value=actor))
    style_body(ws.cell(row=r, column=3, value=atype[1]))
    style_body(ws.cell(row=r, column=4, value=atype[4]))
    style_body(ws.cell(row=r, column=5, value=alignment))
    style_body(ws.cell(row=r, column=6, value=contribution))
    style_body(ws.cell(row=r, column=7, value=capacity))
    style_body(ws.cell(row=r, column=8, value=f"=E{r}+F{r}+G{r}"))
    id_f, name_f, colour_f = band_formulas("H" + str(r), "I", r, "Value")
    style_body(ws.cell(row=r, column=9, value=id_f))
    style_body(ws.cell(row=r, column=10, value=name_f))
    style_body(ws.cell(row=r, column=11, value=colour_f))
    r += 1
last_assistance = r - 1
tab = Table(displayName="Assistance", ref=f"A1:K{last_assistance}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = "A2"

# ---------------------------------------------------------------- Conditional formatting: colour by band
# Colour-code the label cell using the 5 known band colours per RatingType
# (CellIsRule against the *_Name text keeps this readable without VBA).
RISK_BANDS = [("Minimal", "FF63BE7B"), ("Low", "FFA9D18E"), ("Moderate", "FFFFEB84"),
              ("High", "FFF4B183"), ("Very High", "FFE06666")]
VALUE_BANDS = [("Minimal", "FFE06666"), ("Limited", "FFF4B183"), ("Moderate", "FFFFEB84"),
               ("Significant", "FFA9D18E"), ("Exceptional", "FF63BE7B")]

ws_tac = wb["Tactical"]
for band_name, colour in RISK_BANDS:
    ws_tac.conditional_formatting.add(
        f"K2:K{last_tactical}",
        CellIsRule(operator="equal", formula=[f'"{band_name}"'],
                   fill=PatternFill("solid", fgColor=colour, bgColor=colour))
    )

for sheet_name, last_row in (("Initiative", last_initiative), ("Assistance", last_assistance)):
    ws_v = wb[sheet_name]
    for band_name, colour in VALUE_BANDS:
        ws_v.conditional_formatting.add(
            f"J2:J{last_row}",
            CellIsRule(operator="equal", formula=[f'"{band_name}"'],
                       fill=PatternFill("solid", fgColor=colour, bgColor=colour))
        )

wb.move_sheet("Instructions", offset=-10)
wb.active = 0

wb.save(sys.argv[1] if len(sys.argv) > 1 else "management-preparation.xlsx")
print("saved", sys.argv[1] if len(sys.argv) > 1 else "management-preparation.xlsx")
print("tactical rows:", last_tactical - 1, "initiative rows:", last_initiative - 1,
      "assistance rows:", last_assistance - 1)
