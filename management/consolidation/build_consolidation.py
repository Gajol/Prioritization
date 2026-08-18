"""
Generate the Management-side Consolidation workbook shell: an Instructions
sheet (incl. the source-folder config Power Query reads), plus read-only
copies of the reference tables the combined per-centre data needs to relate
to (Priority, Resources, Centres) — the same tables/columns as
preparation.xlsx and centre-template.xlsx, reusing their table-building
helpers directly rather than re-deriving the pattern.

What this script deliberately does NOT do: create the Teams_All/
Priorities_All/ResourceAllocation_All sheets. Those are Power Query query
outputs, authored once by hand in Excel's Query Editor (see "Consolidation
workbook" in docs/excel-file-design.md for the recipe) and loaded as Tables
by Excel itself — openpyxl pre-creating placeholder sheets/Tables with the
same names would just create a collision for Power Query's own "Load To"
step to sort out. openpyxl's job here is only the parts it can safely own:
static structure that doesn't change on refresh.

Usage:
    python build_consolidation.py <preparation.xlsx> <output-consolidation.xlsx> [default-source-folder]

After running:
    1. Open the output file in Excel.
    2. Author the 3 Folder-connector Power Query queries by hand (recipe in
       docs/excel-file-design.md) against the folder named on the
       Instructions sheet.
    3. Run wire_data_model.py (this directory) to wire the combined tables
       and reference tables into the Power Pivot Data Model.
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "templates" / "scripts"))
from build_centre_template import (  # noqa: E402
    FONT, style_header, style_body, read_table, write_reference_table,
)


def main(prep_path, out_path, default_source_folder=""):
    src = openpyxl.load_workbook(prep_path, data_only=True)
    wb = openpyxl.Workbook()

    # ================================================================= Instructions
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    text_blocks = [
        ("Prioritization — Consolidation Workbook", 14, True),
        ("", None, False),
        ("Purpose", 12, True),
        ("Combines the completed centre-template.xlsx files returned by all "
         "six Centre Leads into consolidated views: which priorities each "
         "centre is working, how they're ranked, and how much resourcing "
         "(FTE) each priority is actually getting.", None, False),
        ("", None, False),
        ("Setup", 12, True),
        ("1. Put all six returned centre files in one folder — the yellow "
         "cell below points at it. A local folder or a SharePoint-synced "
         "folder both work; a live SharePoint connection is deliberately "
         "not used here (see docs/excel-file-design.md for why).", None, False),
        ("2. The three combined queries (Teams_All, Priorities_All, "
         "ResourceAllocation_All) and the Data Model relationships/measures "
         "are authored once — see docs/excel-file-design.md's "
         "'Consolidation workbook' section for the exact recipe. This "
         "workbook ships with just the shell and the static reference "
         "tables below; the rest is a one-time Excel UI setup step.", None, False),
        ("3. Every quarter: refresh the returned files into the source "
         "folder, then Data -> Refresh All.", None, False),
        ("", None, False),
        ("Why Team Name alone isn't enough", 12, True),
        ("Centres can name teams however they like — two different centres "
         "may both have a team called the same thing. The combined tables "
         "carry a TeamKey column (Centre Code | Team Name) precisely to "
         "keep those separate; don't rely on Team Name alone when building "
         "anything new here.", None, False),
        ("", None, False),
        ("Source folder", 12, True),
    ]
    r = 1
    for text, size, bold in text_blocks:
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=FONT, bold=bool(bold), size=size or 11)
        if text and not bold and size is None:
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 30
        r += 1

    # A one-row Table, not a plain cell: Power Query's standard parameter
    # pattern is Excel.CurrentWorkbook(){[Name="Config"]}[Content]{0}[SourceFolder]
    # — a named Table survives being moved around the sheet; a hardcoded
    # cell address doesn't.
    config_header_row = r
    style_header(ws.cell(row=config_header_row, column=1, value="SourceFolder"))
    style_body(ws.cell(row=config_header_row + 1, column=1, value=default_source_folder), editable=True)
    ws.column_dimensions["A"].width = 100
    tab = Table(displayName="Config", ref=f"A{config_header_row}:A{config_header_row + 1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)

    # ================================================================= Reference sheets (read-only)
    def copy_ref(sheet_name, col_widths):
        headers, rows = read_table(src, sheet_name)
        ref_ws = wb.create_sheet(sheet_name)
        write_reference_table(ref_ws, sheet_name, 1, headers, rows, col_widths)

    copy_ref("Centres", [6, 26, 12])

    pr_headers, pr_rows = read_table(src, "Priority")
    ws = wb.create_sheet("Priority")
    write_reference_table(ws, "Priority", 1, pr_headers, pr_rows, [42, 12, 10, 12])

    res_headers, res_rows = read_table(src, "Resources")
    ws = wb.create_sheet("Resources")
    for i, h in enumerate(res_headers):
        style_header(ws.cell(row=1, column=1 + i, value=h))
    style_header(ws.cell(row=1, column=len(res_headers) + 1, value="FullName"))
    for r_off, row in enumerate(res_rows):
        row_num = 2 + r_off
        for c_off, val in enumerate(row):
            style_body(ws.cell(row=row_num, column=1 + c_off, value=val))
        style_body(ws.cell(row=row_num, column=len(res_headers) + 1,
                            value=f"=A{row_num}&\" \"&B{row_num}"))
    last_res_row = 1 + len(res_rows)
    tab = Table(displayName="Resources", ref=f"A1:D{last_res_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)
    ws.add_table(tab)
    for col, w in zip("ABCD", [16, 16, 26, 26]):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True
    wb.defined_names["ResourceNameList"] = DefinedName(
        "ResourceNameList", attr_text=f"Resources!$D$2:$D${last_res_row}")

    wb.move_sheet("Instructions", offset=-10)
    wb.active = 0
    wb.save(out_path)
    print("saved", out_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        raise SystemExit(
            "Usage: python build_consolidation.py <preparation.xlsx> <output.xlsx> [default-source-folder]"
        )
    main(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else "")
